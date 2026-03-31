from __future__ import annotations

from pathlib import Path
from datetime import datetime, timezone
import xml.etree.ElementTree as ET
import zipfile

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore[import-untyped]

from xlinject.cli_write_cells import _load_cells_from_csv, _load_cells_from_json, _load_cells_from_json_text
from xlinject.highlevel import apply_recalc_policy, build_column_cell_map, inject_cells, inject_cells_mixed, merge_cell_maps, normalize_numeric_value, to_excel_serial
from xlinject.injector import replace_sentinel_in_column_range, validate_cell_values, write_cells, write_numeric_cells
from xlinject.workbook_map import map_sheet_name_to_part

NS = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
NS_CT = {"ct": "http://schemas.openxmlformats.org/package/2006/content-types"}
NS_REL = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}


def _create_test_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("No active worksheet")
    ws.title = "Eingabemaske"

    ws["A1"] = "Timestamp"
    ws["B1"] = "Wert"

    ws["B2"] = -1
    ws["B3"] = 42
    ws["B4"] = -1
    ws["B5"] = -1
    ws["B6"] = -1

    ws["H2"] = (
        "=LET("
        "werte,B2:B6,"
        "valide,(ISTZAHL(werte))*(werte<>-1),"
        "res,FILTER(werte,valide,\"\"),"
        "WENN(ODER(res=\"\",ANZAHL(res)=0),\"\",MAX(res))"
        ")"
    )

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _create_mixed_test_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("No active worksheet")
    ws.title = "Template"

    ws["B10"] = "BK4S1-0000001"
    ws["B13"] = "12345678901"
    ws["B14"] = "DE0000000000000000000000000000001"
    ws["B16"] = 100
    ws["B17"] = 80
    ws["B18"] = 1000
    ws["B19"] = '=IF(AND(ISNUMBER(B16),ISNUMBER(B18),B18<>0),B18/B16,"")'
    ws["B20"] = "bitte auswählen"
    ws["B23"] = 100.25
    ws["B24"] = 80.25
    ws["B25"] = "=B23-B24"

    for cell_ref in ["B36", "B42", "B43"]:
        ws[cell_ref].number_format = "@"

    validations = [
        (DataValidation(type="textLength", operator="equal", formula1="13", allow_blank=True), "B10"),
        (DataValidation(type="textLength", operator="equal", formula1="11", allow_blank=True), "B13"),
        (DataValidation(type="textLength", operator="equal", formula1="33", allow_blank=True), "B14"),
        (DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True), "B16:B18"),
        (DataValidation(type="list", formula1='"ja,nein,bitte auswählen"', allow_blank=True), "B20"),
        (DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True), "B23:B24"),
    ]

    for validation, range_ref in validations:
        ws.add_data_validation(validation)
        validation.add(range_ref)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    _patch_cell_to_shared_string(path, "Template", "B10")


def _patch_cell_to_shared_string(xlsx_path: Path, sheet_name: str, cell_ref: str) -> None:
    with zipfile.ZipFile(xlsx_path, "r") as zin:
        sheet_part = map_sheet_name_to_part(zin, sheet_name)
        sheet_root = ET.fromstring(zin.read(sheet_part))
        workbook_rels = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))
        content_types = ET.fromstring(zin.read("[Content_Types].xml"))
        archive_data = {item.filename: zin.read(item.filename) for item in zin.infolist()}

    cell = sheet_root.find(f".//x:c[@r='{cell_ref}']", NS)
    if cell is None:
        raise RuntimeError(f"Cell not found for shared string patch: {cell_ref}")

    text_node = cell.find("x:is/x:t", NS)
    shared_text = "" if text_node is None or text_node.text is None else text_node.text

    is_node = cell.find("x:is", NS)
    if is_node is not None:
        cell.remove(is_node)
    value_node = cell.find("x:v", NS)
    if value_node is None:
        value_node = ET.SubElement(cell, "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v")
    value_node.text = "0"
    cell.attrib["t"] = "s"

    rel_type = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings"
    rel_tag = "{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"
    has_rel = any(rel.attrib.get("Type") == rel_type for rel in workbook_rels.findall(rel_tag))
    if not has_rel:
        existing_ids = [rel.attrib.get("Id", "") for rel in workbook_rels.findall(rel_tag)]
        next_id = len(existing_ids) + 1
        while f"rId{next_id}" in existing_ids:
            next_id += 1
        ET.SubElement(
            workbook_rels,
            rel_tag,
            {
                "Id": f"rId{next_id}",
                "Type": rel_type,
                "Target": "sharedStrings.xml",
            },
        )

    override_tag = "{http://schemas.openxmlformats.org/package/2006/content-types}Override"
    has_override = any(
        node.attrib.get("PartName") == "/xl/sharedStrings.xml"
        for node in content_types.findall(override_tag)
    )
    if not has_override:
        ET.SubElement(
            content_types,
            override_tag,
            {
                "PartName": "/xl/sharedStrings.xml",
                "ContentType": (
                    "application/vnd.openxmlformats-officedocument.spreadsheetml."
                    "sharedStrings+xml"
                ),
            },
        )

    shared_root = ET.Element(
        "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sst",
        {"count": "1", "uniqueCount": "1"},
    )
    si_node = ET.SubElement(shared_root, "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si")
    t_node = ET.SubElement(si_node, "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t")
    t_node.text = shared_text

    archive_data[sheet_part] = ET.tostring(sheet_root, encoding="utf-8", xml_declaration=True)
    archive_data["xl/_rels/workbook.xml.rels"] = ET.tostring(
        workbook_rels,
        encoding="utf-8",
        xml_declaration=True,
    )
    archive_data["[Content_Types].xml"] = ET.tostring(
        content_types,
        encoding="utf-8",
        xml_declaration=True,
    )
    archive_data["xl/sharedStrings.xml"] = ET.tostring(
        shared_root,
        encoding="utf-8",
        xml_declaration=True,
    )

    with zipfile.ZipFile(xlsx_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for filename, data in archive_data.items():
            zout.writestr(filename, data)


def _read_cell_v_text(xlsx_path: Path, cell_ref: str) -> str | None:
    with zipfile.ZipFile(xlsx_path, "r") as archive:
        sheet_part = map_sheet_name_to_part(archive, "Eingabemaske")
        root = ET.fromstring(archive.read(sheet_part))

    cell = root.find(f".//x:c[@r='{cell_ref}']", NS)
    if cell is None:
        return None
    value = cell.find("x:v", NS)
    return value.text if value is not None else None


def _read_formula_text(xlsx_path: Path, cell_ref: str) -> str | None:
    with zipfile.ZipFile(xlsx_path, "r") as archive:
        sheet_part = map_sheet_name_to_part(archive, "Eingabemaske")
        root = ET.fromstring(archive.read(sheet_part))

    cell = root.find(f".//x:c[@r='{cell_ref}']", NS)
    if cell is None:
        return None
    formula = cell.find("x:f", NS)
    return formula.text if formula is not None else None


def _read_cell_xml(xlsx_path: Path, sheet_name: str, cell_ref: str) -> str | None:
    with zipfile.ZipFile(xlsx_path, "r") as archive:
        sheet_part = map_sheet_name_to_part(archive, sheet_name)
        root = ET.fromstring(archive.read(sheet_part))

    cell = root.find(f".//x:c[@r='{cell_ref}']", NS)
    if cell is None:
        return None
    return ET.tostring(cell, encoding="unicode")


def _read_data_validations_xml(xlsx_path: Path, sheet_name: str) -> str | None:
    with zipfile.ZipFile(xlsx_path, "r") as archive:
        sheet_part = map_sheet_name_to_part(archive, sheet_name)
        root = ET.fromstring(archive.read(sheet_part))

    node = root.find("x:dataValidations", NS)
    if node is None:
        return None
    return ET.tostring(node, encoding="unicode")


def _read_cell_inline_or_value(xlsx_path: Path, sheet_name: str, cell_ref: str) -> str | None:
    with zipfile.ZipFile(xlsx_path, "r") as archive:
        sheet_part = map_sheet_name_to_part(archive, sheet_name)
        root = ET.fromstring(archive.read(sheet_part))

    cell = root.find(f".//x:c[@r='{cell_ref}']", NS)
    if cell is None:
        return None
    inline_text = cell.find("x:is/x:t", NS)
    if inline_text is not None:
        return inline_text.text
    value_text = cell.find("x:v", NS)
    return value_text.text if value_text is not None else None


def test_replace_only_sentinel_cells_and_preserve_formula(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _create_test_workbook(source)

    before_formula = _read_formula_text(source, "H2")

    report = replace_sentinel_in_column_range(
        source,
        output,
        sheet_name="Eingabemaske",
        range_ref="B2:B6",
        values=[1.1, 2.2, 3.3],
        sentinel=-1.0,
        guard_cells=["H2"],
    )

    assert report.replaced_count == 3
    assert report.consumed_values == 3
    assert report.untouched_sentinel_count == 1

    assert _read_cell_v_text(output, "B2") == "1.1"
    assert _read_cell_v_text(output, "B3") == "42"
    assert _read_cell_v_text(output, "B4") == "2.2"
    assert _read_cell_v_text(output, "B5") == "3.3"
    assert _read_cell_v_text(output, "B6") == "-1"

    after_formula = _read_formula_text(output, "H2")
    assert before_formula == after_formula


def test_guard_cell_change_raises(tmp_path: Path) -> None:
    source = tmp_path / "source_guard.xlsx"
    output = tmp_path / "output_guard.xlsx"
    _create_test_workbook(source)

    try:
        replace_sentinel_in_column_range(
            source,
            output,
            sheet_name="Eingabemaske",
            range_ref="B2:B2",
            values=[8.8],
            sentinel=-1.0,
            guard_cells=["B2"],
        )
    except RuntimeError as exc:
        assert "Guard cell signatures changed" in str(exc)
    else:
        raise AssertionError("Expected RuntimeError when guard cell is modified")


def test_write_numeric_cells_preserve_formula_and_skip_nan(tmp_path: Path) -> None:
    source = tmp_path / "source_write.xlsx"
    output = tmp_path / "output_write.xlsx"
    _create_test_workbook(source)

    before_formula = _read_formula_text(source, "H2")

    report = write_numeric_cells(
        source,
        output,
        sheet_name="Eingabemaske",
        cell_values={"B2": 5.5, "B4": 6.6, "B7": float("nan")},
        guard_cells=["H2"],
    )

    assert report.written_count == 2
    assert report.skipped_nan_count == 1

    assert _read_cell_v_text(output, "B2") == "5.5"
    assert _read_cell_v_text(output, "B4") == "6.6"

    after_formula = _read_formula_text(output, "H2")
    assert before_formula == after_formula


def test_write_numeric_cells_in_place(tmp_path: Path) -> None:
    source = tmp_path / "source_in_place.xlsx"
    _create_test_workbook(source)

    report = write_numeric_cells(
        source,
        source,
        sheet_name="Eingabemaske",
        cell_values={"B5": 9.9},
    )

    assert report.written_count == 1
    assert _read_cell_v_text(source, "B5") == "9.9"


def test_write_numeric_cells_reject_formula_overwrite(tmp_path: Path) -> None:
    source = tmp_path / "source_formula_guard.xlsx"
    output = tmp_path / "output_formula_guard.xlsx"
    _create_test_workbook(source)

    try:
        write_numeric_cells(
            source,
            output,
            sheet_name="Eingabemaske",
            cell_values={"H2": 12.3},
        )
    except RuntimeError as exc:
        assert "Refusing to overwrite formula cell" in str(exc)
    else:
        raise AssertionError("Expected RuntimeError when attempting to overwrite formula cell")


def test_write_numeric_cells_preserves_ignorable_namespaces(tmp_path: Path) -> None:
    source = tmp_path / "source_ns.xlsx"
    output = tmp_path / "output_ns.xlsx"
    _create_test_workbook(source)

    with zipfile.ZipFile(source, "r") as zin:
        sheet_part = map_sheet_name_to_part(zin, "Eingabemaske")
        xml_text = zin.read(sheet_part).decode("utf-8")

    if "mc:Ignorable" not in xml_text:
        xml_text = xml_text.replace(
            "<worksheet ",
            (
                "<worksheet "
                'xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" '
                'xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac" '
                'xmlns:xr="http://schemas.microsoft.com/office/spreadsheetml/2014/revision" '
                'xmlns:xr2="http://schemas.microsoft.com/office/spreadsheetml/2015/revision2" '
                'xmlns:xr3="http://schemas.microsoft.com/office/spreadsheetml/2016/revision3" '
                'mc:Ignorable="x14ac xr xr2 xr3" '
            ),
            1,
        )

    with zipfile.ZipFile(source, "r") as zin:
        with zipfile.ZipFile(output, "w") as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == sheet_part:
                    data = xml_text.encode("utf-8")
                zout.writestr(item, data)

    report = write_numeric_cells(
        output,
        output,
        sheet_name="Eingabemaske",
        cell_values={"B2": 7.7},
        guard_cells=["H2"],
    )

    assert report.written_count == 1

    with zipfile.ZipFile(output, "r") as z:
        out_head = z.read(sheet_part).decode("utf-8")[:1800]

    assert "Ignorable=\"x14ac xr xr2 xr3\"" in out_head
    assert "xmlns:x14ac=" in out_head
    assert "xmlns:xr=" in out_head
    assert "xmlns:xr2=" in out_head
    assert "xmlns:xr3=" in out_head


def test_load_cells_from_json_mapping(tmp_path: Path) -> None:
    path = tmp_path / "cells.json"
    path.write_text('{"b45": 1.25, "C46": 2}', encoding="utf-8")

    loaded = _load_cells_from_json(path)
    assert loaded == {"B45": 1.25, "C46": 2.0}


def test_load_cells_from_csv(tmp_path: Path) -> None:
    path = tmp_path / "cells.csv"
    path.write_text("cell,value\n b45 ,1.25\nC46,2\n", encoding="utf-8")

    loaded = _load_cells_from_csv(path)
    assert loaded == {"B45": 1.25, "C46": 2.0}


def test_load_cells_from_json_text_mapping() -> None:
    loaded = _load_cells_from_json_text('{"b45": 1.25, "C46": 2}')
    assert loaded == {"B45": 1.25, "C46": 2.0}


def test_inject_cells_highlevel_accepts_str_values(tmp_path: Path) -> None:
    source = tmp_path / "source_highlevel.xlsx"
    output = tmp_path / "output_highlevel.xlsx"
    _create_test_workbook(source)

    report = inject_cells(
        source,
        output,
        sheet_name="Eingabemaske",
        cell_values={"b2": "5,5", "B4": 6.6, "B7": None, "B8": float("nan")},
        guard_cells=["H2"],
    )

    assert report.written_count == 2
    assert _read_cell_v_text(output, "B2") == "5.5"
    assert _read_cell_v_text(output, "B4") == "6.6"


def test_normalize_numeric_value() -> None:
    assert normalize_numeric_value(None) is None
    assert normalize_numeric_value("") is None
    assert normalize_numeric_value(" 12,34 ") == 12.34
    assert normalize_numeric_value(float("nan")) is None
    assert normalize_numeric_value("abc") is None


def test_to_excel_serial_with_datetime_like() -> None:
    naive = datetime(2025, 1, 1, 0, 0, 0)
    aware = datetime(2025, 1, 1, 0, 0, 0, tzinfo=timezone.utc)

    serial_naive = to_excel_serial(naive)
    serial_aware = to_excel_serial(aware)

    assert serial_naive is not None
    assert serial_aware is not None
    assert abs(serial_naive - serial_aware) < 1e-9


def test_build_column_map_and_merge() -> None:
    col_b = build_column_cell_map(2, [1.0, None, "3,5"], 45)
    col_c = build_column_cell_map(3, [10, 20], 45)
    merged = merge_cell_maps(col_b, col_c)

    assert col_b == {"B45": 1.0, "B47": 3.5}
    assert col_c == {"C45": 10.0, "C46": 20.0}
    assert merged["B45"] == 1.0
    assert merged["B47"] == 3.5
    assert merged["C46"] == 20.0


def test_inject_cells_sets_calcpr_recalc_flags(tmp_path: Path) -> None:
    source = tmp_path / "source_recalc.xlsx"
    output = tmp_path / "output_recalc.xlsx"
    _create_test_workbook(source)

    report = inject_cells(
        source,
        output,
        sheet_name="Eingabemaske",
        cell_values={"B2": 5.5},
    )

    assert report.written_count == 1

    with zipfile.ZipFile(output, "r") as z:
        wb_root = ET.fromstring(z.read("xl/workbook.xml"))
        calc_pr = wb_root.find("x:calcPr", NS)
        assert calc_pr is not None
        assert calc_pr.attrib.get("fullCalcOnLoad") == "1"
        assert calc_pr.attrib.get("forceFullCalc") == "1"
        assert calc_pr.attrib.get("calcOnSave") == "1"
        assert calc_pr.attrib.get("calcCompleted") == "0"


def test_apply_recalc_policy_can_clear_formula_cached_values(tmp_path: Path) -> None:
    source = tmp_path / "source_formula_cache.xlsx"
    _create_test_workbook(source)

    # Ensure formula cell has cached value node for the test scenario
    with zipfile.ZipFile(source, "r") as zin:
        sheet_part = map_sheet_name_to_part(zin, "Eingabemaske")
        xml_text = zin.read(sheet_part).decode("utf-8")

    if "<f>LET(" in xml_text and "<c r=\"H2\"" in xml_text and "<v>" not in xml_text:
        xml_text = xml_text.replace("<c r=\"H2\"><f>", "<c r=\"H2\"><f>", 1)
        xml_text = xml_text.replace("</f></c>", "</f><v>999</v></c>", 1)

        with zipfile.ZipFile(source, "r") as zin:
            temp = tmp_path / "patched.xlsx"
            with zipfile.ZipFile(temp, "w") as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if item.filename == sheet_part:
                        data = xml_text.encode("utf-8")
                    zout.writestr(item, data)
        source = temp

    apply_recalc_policy(
        source,
        remove_calc_chain_file=False,
        set_full_calc_on_load=False,
        clear_formula_cached_values=True,
        clear_formula_cache_sheets=("Eingabemaske",),
    )

    with zipfile.ZipFile(source, "r") as z:
        sheet_part = map_sheet_name_to_part(z, "Eingabemaske")
        root = ET.fromstring(z.read(sheet_part))
        formula_cell = root.find(".//x:c[@r='H2']", NS)
        assert formula_cell is not None
        assert formula_cell.find("x:f", NS) is not None
        assert formula_cell.find("x:v", NS) is None


def test_write_cells_supports_mixed_values_and_preserves_neighbors(tmp_path: Path) -> None:
    source = tmp_path / "source_mixed.xlsx"
    output = tmp_path / "output_mixed.xlsx"
    _create_mixed_test_workbook(source)

    before_formula_b19 = _read_cell_xml(source, "Template", "B19")
    before_formula_b25 = _read_cell_xml(source, "Template", "B25")
    before_validations = _read_data_validations_xml(source, "Template")

    report = write_cells(
        source,
        output,
        sheet_name="Template",
        cell_values={
            "B10": "BK4S1-0008738",
            "B13": "51354532108",
            "B14": "DE0006973844010322502900000980011",
            "B16": 310,
            "B17": 129,
            "B18": 522942,
            "B20": "ja",
            "B23": 57501.08,
            "B24": 29851.34,
            "B36": "SITE_A GmbH",
            "B42": "Example Street 1",
            "B43": "12345 Example City",
        },
        guard_cells=["B19", "B25"],
        validate_sheet_rules=True,
    )

    assert report.written_count == 12
    assert _read_cell_inline_or_value(output, "Template", "B10") == "BK4S1-0008738"
    assert _read_cell_inline_or_value(output, "Template", "B14") == "DE0006973844010322502900000980011"
    assert _read_cell_inline_or_value(output, "Template", "B20") == "ja"
    assert _read_cell_inline_or_value(output, "Template", "B36") == "SITE_A GmbH"
    assert _read_cell_inline_or_value(output, "Template", "B42") == "Example Street 1"
    assert _read_cell_inline_or_value(output, "Template", "B43") == "12345 Example City"

    assert 't="inlineStr"' in (_read_cell_xml(output, "Template", "B10") or "")
    assert 't="inlineStr"' in (_read_cell_xml(output, "Template", "B36") or "")
    assert before_formula_b19 == _read_cell_xml(output, "Template", "B19")
    assert before_formula_b25 == _read_cell_xml(output, "Template", "B25")
    assert before_validations == _read_data_validations_xml(output, "Template")


def test_validate_cell_values_rejects_invalid_whole_number(tmp_path: Path) -> None:
    source = tmp_path / "source_validation.xlsx"
    _create_mixed_test_workbook(source)

    try:
        validate_cell_values(
            source,
            sheet_name="Template",
            cell_values={"B17": 129.2, "B20": "ja", "B10": "BK4S1-0008738"},
        )
    except ValueError as exc:
        assert "Cell B17 expects a whole number" in str(exc)
    else:
        raise AssertionError("Expected ValueError for invalid whole-number validation")


def test_inject_cells_mixed_sets_calcpr_recalc_flags(tmp_path: Path) -> None:
    source = tmp_path / "source_mixed_recalc.xlsx"
    output = tmp_path / "output_mixed_recalc.xlsx"
    _create_mixed_test_workbook(source)

    report = inject_cells_mixed(
        source,
        output,
        sheet_name="Template",
        cell_values={
            "B10": "BK4S1-0008738",
            "B16": 310,
            "B17": 129,
            "B18": 522942,
            "B20": "ja",
        },
        guard_cells=["B19", "B25"],
    )

    assert report.written_count == 5

    with zipfile.ZipFile(output, "r") as z:
        wb_root = ET.fromstring(z.read("xl/workbook.xml"))
        calc_pr = wb_root.find("x:calcPr", NS)
        assert calc_pr is not None
        assert calc_pr.attrib.get("fullCalcOnLoad") == "1"
        assert calc_pr.attrib.get("forceFullCalc") == "1"
        assert calc_pr.attrib.get("calcOnSave") == "1"
        assert calc_pr.attrib.get("calcCompleted") == "0"
